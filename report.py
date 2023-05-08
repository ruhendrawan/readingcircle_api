# pip install openpyxl

import ast
import argparse
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


def read_data(file_path, delimiter):
    data = []
    with open(file_path, newline='', encoding='utf-8', errors='replace') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=delimiter)
        for row in reader:
            data.append(row)
    return data


def col_letter(col_idx):
    letter = ""
    while col_idx:
        letter = chr((col_idx - 1) % 26 + 65) + letter
        col_idx = (col_idx - 1) // 26
    return letter


def process_data(data):
    user_docno_correct = defaultdict(lambda: defaultdict(int))
    user_docsrc_correct = defaultdict(lambda: defaultdict(int))
    docno_question_counts = defaultdict(int)
    docsrc_question_counts = defaultdict(int)

    for row in data:
        user_docno_correct[row['usr']][row['docno']] += int(row['correct'])
        user_docsrc_correct[row['usr']][row['docsrc']] += int(row['correct'])

        docno_question_counts[row['docno']] = max(docno_question_counts[row['docno']], user_docno_correct[row['usr']][row['docno']])
        docsrc_question_counts[row['docsrc']] = max(docsrc_question_counts[row['docsrc']], user_docsrc_correct[row['usr']][row['docsrc']])

    return user_docno_correct, user_docsrc_correct, docno_question_counts, docsrc_question_counts

def write_to_excel(user_docno_correct, user_docsrc_correct, docno_question_counts, docsrc_question_counts, output_file):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "User_DocNo"
    header = ["User"] + [f"{docno}" for docno in sorted(docno_question_counts)]
    ws1.append(header)
    max_row = ["Max"] + [docno_question_counts[docno] for docno in sorted(docno_question_counts)]
    ws1.append(max_row)

    for usr, docno_correct in user_docno_correct.items():
        user_row = [usr] + [docno_correct[docno] for docno in sorted(docno_question_counts)]
        ws1.append(user_row)

    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    ws2 = wb.create_sheet("User_DocSrc", -1)
    header = ["User"] + [f"{docno}" for docno in sorted(docsrc_question_counts)]
    ws2.append(header)
    max_row = ["Max"] + [docsrc_question_counts[docno] for docno in sorted(docsrc_question_counts)]
    ws2.append(max_row)
    # Add formula to compute the sum of columns B, C, and D
    cell = ws2.cell(row=2, column=len(header) + 1,
             value=f"=SUM(B{2}:{col_letter(len(header))}{2})")
    cell.fill = gray_fill

    row_number = 3
    for usr, docno_correct in user_docsrc_correct.items():
        user_row = [usr] + [docno_correct[docno] for docno in sorted(docsrc_question_counts)]
        ws2.append(user_row)

        # Add formula to compute the sum of columns B, C, and D
        cell = ws2.cell(row=row_number, column=len(user_row) + 1,
                 value=f"=SUM(B{row_number}:{col_letter(len(header))}{row_number})")
        cell.fill = gray_fill

        # Add formula to compute column E divided by $E$2
        cell = ws2.cell(row=row_number, column=len(user_row) + 2,
                 value=f"=ROUND({col_letter(len(header) + 1)}{row_number}/${col_letter(len(header) + 1)}$2*100, 2)")
        cell.font = Font(bold=True)
        cell.fill = gray_fill

        row_number += 1

    wb.save(output_file)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process quiz data and generate a report.")
    parser.add_argument("input_file", help="Path to the input TSV or CSV file containing quiz data.")
    parser.add_argument("output_file", help="Path where the generated Excel file will be saved.")
    parser.add_argument("-d", "--delimiter", default=',', help="Delimiter for the input file. Use '\\t' for TSV files and ',' for CSV files. Default: '\\t'")

    args = parser.parse_args()

    if args.delimiter == '\\t':
        data = read_data(args.input_file, '\t')
    else:
        data = read_data(args.input_file, args.delimiter)
    user_docno_correct, user_docsrc_correct, docno_question_counts, docsrc_question_counts = process_data(data)

    write_to_excel(user_docno_correct, user_docsrc_correct, docno_question_counts, docsrc_question_counts, args.output_file)
    print(f"Results saved to {args.output_file}")
