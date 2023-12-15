# pip install openpyxl

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


def col_letter(col_idx):
    letter = ""
    while col_idx > 0:
        col_idx -= 1
        letter = chr(col_idx % 26 + 65) + letter
        col_idx //= 26
    return letter



def process_quiz_results(data):
    user_docno_correct = defaultdict(lambda: defaultdict(int))
    user_docsrc_correct = defaultdict(lambda: defaultdict(int))
    docno_question_counts = defaultdict(int)
    docsrc_question_counts = defaultdict(int)

    for row in data:
        user_docno_correct[row['usr']][row['docno']] += int(row['correct'])
        user_docsrc_correct[row['usr']][row['docsrc']] += int(row['correct'])

        docno_question_counts[row['docno']] = \
            max(docno_question_counts[row['docno']],
                user_docno_correct[row['usr']][row['docno']])
        docsrc_question_counts[row['docsrc']] = \
            max(docsrc_question_counts[row['docsrc']],
                user_docsrc_correct[row['usr']][row['docsrc']])

    return \
        user_docno_correct, user_docsrc_correct, \
            docno_question_counts, docsrc_question_counts

def write_quiz_results_to_excel(
        user_docno_correct, user_docsrc_correct,
        docno_question_counts, docsrc_question_counts,
        output_file):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "User_DocNo"
    header = ["User"] + ["{0}".format(docno) for docno in sorted(docno_question_counts)]
    ws1.append(header)
    max_row = ["Max"] + [docno_question_counts[docno] for docno in sorted(docno_question_counts)]
    ws1.append(max_row)

    for usr, docno_correct in user_docno_correct.items():
        user_row = [usr] + [docno_correct[docno] for docno in sorted(docno_question_counts)]
        ws1.append(user_row)

    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    ws2 = wb.create_sheet("User_DocSrc", -1)
    header = ["User"] + ["{0}".format(docno) for docno in sorted(docsrc_question_counts)]
    ws2.append(header)
    max_row = ["Max"] + [docsrc_question_counts[docno] for docno in sorted(docsrc_question_counts)]
    ws2.append(max_row)
    # Add formula to compute the sum of columns B, C, and D
    # cell = ws2.cell(row=2, column=len(header) + 1,
    #          value="=SUM(B{0}:{col_letter(len(header))}{0})".format(2))
    # cell.fill = gray_fill

    row_number = 3
    for usr, docno_correct in user_docsrc_correct.items():
        user_row = [usr] + [docno_correct[docno] for docno in sorted(docsrc_question_counts)]
        ws2.append(user_row)

        # Add formula to compute the sum of columns B, C, and D
        # cell = ws2.cell(row=row_number, column=len(user_row) + 1,
        #          value="=SUM(B{0}:{col_letter(len(header))}{0})".format(row_number))
        # cell.fill = gray_fill

        # Add formula to compute column E divided by $E$2
        # cell = ws2.cell(row=row_number, column=len(user_row) + 2,
        #          value="=ROUND({0}{1}/${0}$2*100, 2)".format(col_letter(len(header) + 1),row_number))
        # cell.font = Font(bold=True)
        # cell.fill = gray_fill

        row_number += 1

    wb.save(output_file)
